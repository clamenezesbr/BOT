"""
extractor_quantidades.py
------------------------
Módulo responsável por:
  1. Carregar o mapeamento de produtos (output/mapeamento_produtos.csv).
  2. Ler o texto bruto de cada PDF em temp_pdfs/ e extrair:
       • Número do Pedido de Compra (PO)
       • Pares (ID_Produto, Quantidade)
  3. Deduplicar registros repetidos do rodapé dos PDFs.
  4. Consolidar tudo em um DataFrame Pandas com pivot_table.
  5. Exportar o relatório final estilizado em output/consolidado_pedidos.xlsx.

Layout do Excel gerado:
  Coluna A  → PRODUTO (nome do produto)
  Coluna B  → TOTAL   (soma de todos os POs para aquele produto)
  Colunas C+ → um PO por coluna (ordenados)
"""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from tqdm.contrib.logging import logging_redirect_tqdm

logger = logging.getLogger(__name__)

# ─── Caminhos padrão ────────────────────────────────────────────────────────
PDF_DIR = Path("temp_pdfs")
OUTPUT_DIR = Path("output")
MAPPING_CSV = OUTPUT_DIR / "mapeamento_produtos.csv"
OUTPUT_XLSX = OUTPUT_DIR / "consolidado_pedidos.xlsx"

# ─── Expressões regulares ────────────────────────────────────────────────────
RE_PO_NUMBER = re.compile(r"N[ºo°]\s*do\s*Pedido[:\s]*(\d+)", re.IGNORECASE)

# Captura: ID exato de 7 dígitos (não parte de IDs maiores como 102111257)
# ... texto intermediário ... data dd.mm.aaaa ... quantidade ... UN
# \s+ entre data e quantidade cobre tanto espaço simples quanto quebras de linha
# (fitz pode extrair colunas em linhas separadas dependendo do PDF)
RE_QUANTITY = re.compile(
    r"(?<!\d)(\d{7})(?!\d)[\s\S]{10,300}?(?:\d{2}\.\d{2}\.\d{4})\s+(\d+)\s+U[A-Z]{1,2}\b",
    re.IGNORECASE,
)


# ─── Funções auxiliares ──────────────────────────────────────────────────────

def _load_mapping(csv_path: Path) -> dict[str, str]:
    """Carrega o CSV de mapeamento e retorna {ID_PRODUTO: NOME_PRODUTO}."""
    mapping: dict[str, str] = {}
    if not csv_path.exists():
        logger.error("Arquivo de mapeamento não encontrado: '%s'.", csv_path)
        return mapping

    with csv_path.open(encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            pid = row.get("ID_PRODUTO", "").strip()
            pname = row.get("NOME_PRODUTO", "").strip()
            if pid and pname:
                mapping[pid] = pname

    logger.info("Mapeamento carregado: %d produto(s).", len(mapping))
    return mapping


def _extract_text(pdf_path: Path) -> str:
    """Extrai e retorna o texto bruto de todas as páginas de um PDF."""
    try:
        doc = fitz.open(str(pdf_path))
        text = "\n".join(page.get_text("text") for page in doc)
        doc.close()
        return text
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao ler PDF '%s': %s", pdf_path.name, exc)
        return ""


def _parse_pdf(text: str) -> tuple[str | None, dict[str, int]]:
    """
    Extrai o número do PO e o dicionário {ID: quantidade} de um texto de PDF.

    Deduplicação: se o mesmo par (ID, quantidade) aparecer mais de uma vez
    (rodapé repetido), conta apenas uma ocorrência.
    """
    po_match = RE_PO_NUMBER.search(text)
    po_number = po_match.group(1) if po_match else None

    seen: set[tuple[str, int]] = set()
    quantities: dict[str, int] = {}

    for match in RE_QUANTITY.finditer(text):
        prod_id = match.group(1).strip()
        qty_raw = match.group(2).strip()

        try:
            qty = int(qty_raw)
        except ValueError:
            continue

        pair = (prod_id, qty)
        if pair in seen:
            continue
        seen.add(pair)

        # Se o mesmo ID aparecer com quantidades diferentes, soma (improvável,
        # mas defensivo para PDFs com múltiplos itens agrupados por ID).
        quantities[prod_id] = quantities.get(prod_id, 0) + qty

    return po_number, quantities


# ─── Estilo do Excel ─────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_TOTAL_FILL = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_THIN = Side(style="thin", color="B8CCE4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_excel_style(xlsx_path: Path, po_columns: list[str]) -> None:
    """Aplica larguras, bordas, cores e fontes ao arquivo Excel gerado."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # ── Cabeçalho (linha 1) ──────────────────────────────────────────────────
    for cell in ws[1]:
        if cell.column == 2:  # Coluna TOTAL
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER

    # ── Dados (linhas 2+) ────────────────────────────────────────────────────
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        for cell in row:
            cell.border = _BORDER
            if cell.column == 1:
                cell.alignment = _LEFT
                cell.font = _BOLD_FONT
                cell.fill = fill
            elif cell.column == 2:  # TOTAL
                cell.alignment = _CENTER
                cell.font = _BOLD_FONT
                cell.fill = _TOTAL_FILL
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            else:
                cell.alignment = _CENTER
                cell.font = _NORMAL_FONT
                cell.fill = fill

    # ── Larguras das colunas ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 55  # PRODUTO
    ws.column_dimensions["B"].width = 10  # TOTAL
    for col_idx in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # ── Altura do cabeçalho ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30

    # ── Congelar painel no cabeçalho ─────────────────────────────────────────
    ws.freeze_panes = "C2"

    wb.save(xlsx_path)


# ─── Função principal ─────────────────────────────────────────────────────────

def build_consolidation_report(
    pdf_dir: Path = PDF_DIR,
    mapping_csv: Path = MAPPING_CSV,
    output_xlsx: Path = OUTPUT_XLSX,
) -> None:
    """
    Orquestra a leitura dos PDFs, montagem do DataFrame e exportação do Excel.
    """
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    product_map = _load_mapping(mapping_csv)
    if not product_map:
        logger.error("Mapeamento de produtos vazio. Abortando.")
        return

    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    if not pdf_files:
        logger.warning("Nenhum PDF encontrado em '%s'.", pdf_dir)
        return

    records: list[dict] = []

    with logging_redirect_tqdm():
        for pdf_path in tqdm(pdf_files, desc="  📊  Extraindo quantidades", unit="pdf", dynamic_ncols=True, leave=True):
            text = _extract_text(pdf_path)
            if not text:
                continue

            po_number, quantities = _parse_pdf(text)
            if not po_number:
                logger.warning("Número do PO não encontrado em '%s'. Ignorando.", pdf_path.name)
                continue

            if not quantities:
                logger.warning("Nenhuma quantidade extraída de '%s'.", pdf_path.name)
                continue

            for prod_id, qty in quantities.items():
                prod_name = product_map.get(prod_id, f"PRODUTO DESCONHECIDO ({prod_id})")
                records.append(
                    {
                        "PO": po_number,
                        "ID_Produto": prod_id,
                        "Nome_Produto": prod_name,
                        "Quantidade": qty,
                    }
                )

            logger.info("  ↳ PO %s — %d item(ns) extraído(s).", po_number, len(quantities))

    if not records:
        logger.error("Nenhum registro válido foi extraído. Relatório não gerado.")
        return

    df = pd.DataFrame(records)

    # ── Pivot: produtos nas linhas, POs nas colunas ──────────────────────────
    pivot = df.pivot_table(
        index="Nome_Produto",
        columns="PO",
        values="Quantidade",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    pivot.columns.name = None

    # Garante que TODOS os produtos do mapeamento apareçam (mesmo com qty 0)
    all_names = pd.DataFrame(
        {"Nome_Produto": list(product_map.values())}
    ).drop_duplicates()
    pivot = all_names.merge(pivot, on="Nome_Produto", how="left").fillna(0)

    # Ordena por nome do produto
    po_cols = [c for c in pivot.columns if c != "Nome_Produto"]
    po_cols_sorted = sorted(po_cols, key=lambda x: str(x))

    # Coluna TOTAL (soma horizontal)
    pivot["TOTAL"] = pivot[po_cols_sorted].sum(axis=1).astype(int)

    # Layout final: PRODUTO | TOTAL | PO1 | PO2 | ...
    final_df = pivot[["Nome_Produto", "TOTAL"] + po_cols_sorted].copy()
    final_df = final_df.rename(columns={"Nome_Produto": "PRODUTO"})

    # Converte colunas de PO para int
    for col in po_cols_sorted:
        final_df[col] = final_df[col].astype(int)

    # ── Exportação ────────────────────────────────────────────────────────────
    final_df.to_excel(output_xlsx, index=False, engine="openpyxl")
    _apply_excel_style(output_xlsx, [str(c) for c in po_cols_sorted])

    logger.info(
        "Relatório salvo: '%s' (%d produto(s), %d PO(s)).",
        output_xlsx,
        len(final_df),
        len(po_cols_sorted),
    )


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    build_consolidation_report()
    print(f"\n✅  Relatório exportado: '{OUTPUT_XLSX}'.")
