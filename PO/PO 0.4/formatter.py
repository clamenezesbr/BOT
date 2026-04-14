"""
formatter.py
------------
Módulo responsável por:
  - Receber a lista consolidada de registros [{'PO', 'Produto', 'Qtd'}]
  - Detectar e tratar POs duplicados (somando quantidades)
  - Construir uma tabela pivotada (Produto × PO) com coluna TOTAL
  - Garantir que todos os produtos da lista estrita apareçam (Qtd 0 se ausente)
  - Exportar para output/consolidado_pos.xlsx com formatação profissional
"""

import logging
from pathlib import Path
from typing import List, Dict, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from po_parser import PRODUTOS

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paleta de cores
# ---------------------------------------------------------------------------
COLOR_HEADER_BG = "1F4E79"    # Azul escuro — cabeçalho PO / TOTAL
COLOR_HEADER_FG = "FFFFFF"    # Branco — texto do cabeçalho
COLOR_PRODUTO_BG = "D6E4F0"   # Azul claro — coluna Produto
COLOR_TOTAL_BG = "FCE4D6"     # Laranja claro — coluna TOTAL
COLOR_ALT_ROW = "EBF5FB"      # Azul bem claro — linhas alternadas
COLOR_BORDER = "BDC3C7"       # Cinza claro — bordas


def _detect_duplicates(
    records: List[Dict],
) -> Tuple[List[Dict], List[str]]:
    """
    Verifica se há POs duplicados (mesmo PO em múltiplos PDFs).
    Quando encontrado, soma as quantidades e registra no log.

    Returns
    -------
    deduped : list[dict]
        Registros com quantidades somadas por (PO, Produto).
    duplicate_pos : list[str]
        Lista dos POs que foram identificados como duplicatas.
    """
    seen: Dict[Tuple[str, str], int] = {}
    duplicate_pos: List[str] = []

    for rec in records:
        key = (rec["PO"], rec["Produto"])
        if key in seen:
            if rec["PO"] not in duplicate_pos:
                duplicate_pos.append(rec["PO"])
                logger.warning(
                    "PO duplicado detectado: %s — quantidades serão somadas.",
                    rec["PO"],
                )
            seen[key] += rec["Qtd"]
        else:
            seen[key] = rec["Qtd"]

    deduped = [
        {"PO": po, "Produto": produto, "Qtd": qtd}
        for (po, produto), qtd in seen.items()
    ]
    return deduped, duplicate_pos


def build_pivot(records: List[Dict]) -> Tuple[pd.DataFrame, List[str]]:
    """
    Constrói o DataFrame pivotado.

    Estrutura final:
        Produto | TOTAL | <PO_1> | <PO_2> | ...

    - Todos os produtos da lista estrita aparecem (mesmo com Qtd 0).
    - Valores NaN → 0 (int).
    - TOTAL = soma por linha.

    Returns
    -------
    pivot_df : pd.DataFrame
    duplicate_pos : list[str]
        POs duplicados para reporte no log/resumo.
    """
    deduped, duplicate_pos = _detect_duplicates(records)

    if not deduped:
        logger.warning("Nenhum registro válido para construir a tabela.")
        # Retorna DataFrame vazio mas com a estrutura correta
        df = pd.DataFrame({"Produto": PRODUTOS, "TOTAL": [0] * len(PRODUTOS)})
        return df, duplicate_pos

    raw_df = pd.DataFrame(deduped)

    pivot = raw_df.pivot_table(
        index="Produto",
        columns="PO",
        values="Qtd",
        aggfunc="sum",
        fill_value=0,
    )
    pivot.columns.name = None
    pivot.index.name = None

    # Reindexar para garantir todos os produtos (ordem da lista estrita)
    pivot = pivot.reindex(PRODUTOS, fill_value=0)
    pivot = pivot.fillna(0).astype(int)

    # Inserir coluna TOTAL como primeira coluna numérica
    pivot.insert(0, "TOTAL", pivot.sum(axis=1))

    # Transformar index em coluna
    pivot.reset_index(inplace=True)
    pivot.rename(columns={"index": "Produto"}, inplace=True)

    logger.info(
        "Pivot construída: %d produtos × %d POs.",
        len(pivot),
        len(pivot.columns) - 2,  # exclui Produto e TOTAL
    )
    return pivot, duplicate_pos


def _apply_excel_formatting(output_path: Path, n_pos: int) -> None:
    """
    Aplica formatação profissional ao arquivo Excel gerado pelo pandas.

    Parameters
    ----------
    output_path : Path
        Caminho do arquivo .xlsx já criado.
    n_pos : int
        Número de colunas de PO (excluindo Produto e TOTAL).
    """
    wb = load_workbook(str(output_path))
    ws = wb.active
    ws.title = "Consolidado POs"

    total_cols = ws.max_column
    total_rows = ws.max_row

    # ------------------------------------------------------------------
    # Estilos reutilizáveis
    # ------------------------------------------------------------------
    thin_side = Side(style="thin", color=COLOR_BORDER)
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side,
    )

    header_font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
    data_font = Font(size=10)
    produto_font = Font(bold=True, size=10)

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    produto_fill = PatternFill("solid", fgColor=COLOR_PRODUTO_BG)
    total_fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
    alt_fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ------------------------------------------------------------------
    # Linha de cabeçalho (linha 1)
    # ------------------------------------------------------------------
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ------------------------------------------------------------------
    # Linhas de dados
    # ------------------------------------------------------------------
    for row_idx in range(2, total_rows + 1):
        is_alt = (row_idx % 2 == 0)

        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            if col_idx == 1:
                # Coluna Produto
                cell.font = produto_font
                cell.fill = produto_fill
                cell.alignment = left_align
            elif col_idx == 2:
                # Coluna TOTAL
                cell.font = Font(bold=True, size=10)
                cell.fill = total_fill
                cell.alignment = center_align
            else:
                # Colunas de PO
                cell.font = data_font
                cell.fill = alt_fill if is_alt else PatternFill()
                cell.alignment = center_align

    # ------------------------------------------------------------------
    # Largura das colunas
    # ------------------------------------------------------------------
    # Coluna Produto: largura fixa generosa
    ws.column_dimensions[get_column_letter(1)].width = 42
    # Coluna TOTAL
    ws.column_dimensions[get_column_letter(2)].width = 10
    # Colunas de PO
    for col_idx in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # ------------------------------------------------------------------
    # Altura das linhas
    # ------------------------------------------------------------------
    ws.row_dimensions[1].height = 32  # cabeçalho
    for row_idx in range(2, total_rows + 1):
        ws.row_dimensions[row_idx].height = 18

    # ------------------------------------------------------------------
    # Congelar painel: mantém Produto + cabeçalho visíveis ao rolar
    # ------------------------------------------------------------------
    ws.freeze_panes = "C2"

    # ------------------------------------------------------------------
    # Linha de totais no rodapé
    # ------------------------------------------------------------------
    footer_row = total_rows + 1
    ws.cell(row=footer_row, column=1).value = "TOTAL GERAL"
    ws.cell(row=footer_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=footer_row, column=1).fill = PatternFill(
        "solid", fgColor=COLOR_HEADER_BG
    )
    ws.cell(row=footer_row, column=1).font = Font(
        bold=True, color=COLOR_HEADER_FG, size=10
    )
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border

    for col_idx in range(2, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=footer_row, column=col_idx)
        cell.value = f"=SUM({col_letter}2:{col_letter}{total_rows})"
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = center_align
        cell.border = thin_border

    wb.save(str(output_path))
    logger.info("Formatação Excel aplicada com sucesso.")


def export_to_excel(
    records: List[Dict],
    output_dir: Path,
    filename: str = "consolidado_pos.xlsx",
) -> Tuple[Path, List[str]]:
    """
    Ponto de entrada do módulo: constrói a pivot e exporta o Excel formatado.

    Parameters
    ----------
    records : list[dict]
        Registros consolidados {'PO', 'Produto', 'Qtd'}.
    output_dir : Path
        Diretório de saída.
    filename : str
        Nome do arquivo Excel gerado.

    Returns
    -------
    output_path : Path
        Caminho do arquivo gerado.
    duplicate_pos : list[str]
        POs duplicados detectados durante o processo.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / filename

    pivot_df, duplicate_pos = build_pivot(records)

    pivot_df.to_excel(str(output_path), index=False, engine="openpyxl")
    logger.info("Excel salvo em: %s", output_path)

    n_pos = len(pivot_df.columns) - 2  # exclui Produto e TOTAL
    _apply_excel_formatting(output_path, n_pos)

    return output_path, duplicate_pos
